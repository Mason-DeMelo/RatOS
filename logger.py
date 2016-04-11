import openpyxl

class Logger():
	def __init__(self):
		self.fileLocation = "data/"
		self.log = self.fileLocation + "ratLog.xlsx"
		self.detailedLog = list()

	def addToLog(self, side, time):
		self.detailedLog.append([side, time])

	def updateLog(self, number, date, duration, runs, comment, experimenter):
		#Update Log Book
		if number == "" or experimenter == "":
			raise RuntimeError("The form was incomplete!")
		
		try:
			wb = openpyxl.load_workbook(self.log)
		except:
			wb = openpyxl.Workbook()

		wsTitle = "Rat "+str(number)

		if wsTitle in wb:
			ws = wb[wsTitle]
		else:
			ws = wb.create_sheet()
			ws.title = wsTitle
			ws['A1'] = "Date"
			ws['B1'] = "Duration (sec)"
			ws['C1'] = "Runs"
			ws['D1'] = "Comments"
			ws['E1'] = "Experimienter"
			ws['F1'] = "Side"
			ws['G1'] = "Time"

                for event in self.detailedLog:
			row = str(ws.max_row + 1)
                        ws['A'+row] = date
                        ws['B'+row] = duration
                        ws['C'+row] = runs
                        ws['D'+row] = comment
                        ws['E'+row] = experimenter
			ws['F'+row] = event[0]
			ws['G'+row] = event[1]

		wb.save(self.log)
		self.detaileLog = list()
